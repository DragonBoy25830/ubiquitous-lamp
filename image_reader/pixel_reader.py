import re
import urllib.request
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors
import os

input_method = input('Would you like to import a local image or use the image url? [1/2]: ')

if input_method == '2':
    # Checks if url follows established syntax rules to determine validity
    def is_valid_url(url):
        # Source code from django url validation regex --> prevents need to setup django environment
        regex = re.compile(
            r'^(?:http|ftp)s?://' # http:// or https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' #domain...
            r'localhost|' #localhost...
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # ...or ip
            r'(?::\d+)?' # optional port
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)

        return re.match(regex, url) is not None # True

    # Checks if data returned from url is an image
    def is_image(url):
        valid_image_formats = ('image/png', 'image.jpg')
        try:
            site = urllib.request.urlopen(url)
        except:
            return False
        meta = site.info()  # get header of the http request
        if meta["content-type"] in valid_image_formats:  # check if the content-type is a image
            return True
        return False

    # Function that combines results of is_image and is_valid_url functions
    def is_url_ready(url):
        return is_image(url) and is_valid_url(url)

    input_url = input('Enter the URL for the image you would like to convert: ')

    while not is_url_ready(input_url):
        input_url = input('''You entered an invalid url or the url is not going to an image. Please check your url and make sure it's a valid url that goes to an image: ''') 

    urllib.request.urlretrieve(input_url, "input_image.png")

    im = Image.open('input_image.png', 'r')

else:
    global filename
    filename = input('Enter the name of the file: ')

    readmode = True
    while readmode:
        try:
            im = Image.open(filename, 'r')
            readmode = False
        except:
            print(f'Can not read {filename}')
            filename = input('Enter the name of the file: ')

print('\nI would recommend you scale the image down by either the width or the height')
print('Please input only one of the values to still maintain the original image\'s scale and aspect ratio')
print('Just hit enter if you don\'t want to put in a value \n')
print('If you put in no values, the script might take a while to run \n')

width = input("Enter the scalable width (pixels) you would like: ")
height = input("Enter the scalable height (pixels) you would like: ")


if width != '':
    base_width = int(width)
    wpercent = (base_width / float(im.size[0]))
    hsize = int((float(im.size[1]) * float(wpercent)))
    img = im.resize((base_width, hsize), Image.ANTIALIAS)
    img.save('input_resized.png')
elif height != '':
    base_height = int(height)
    hpercent = (base_height / float(im.size[1]))
    wsize = int((float(img.size[0]) * float(hpercent)))
    img = im.resize((wsize, base_height), Image.ANTIALIAS)
    img.save('input_resized.png')

im.close()

img = Image.open('input_resized.png', 'r')
width, height = img.size

pix_val = img.getdata()

def rgb_to_hex(rgb):
    return '#{:02x}{:02x}{:02x}'.format(rgb[0], rgb[1], rgb[2])


workbook = Workbook()
sheet = workbook.active

counter = 0

for i in range(1, height + 1):
    for j in range(1, width + 1):
        color = pix_val[counter]
        counter += 1
        background = PatternFill(fgColor = '00' + rgb_to_hex(color)[1:], fill_type='solid')
        cell = sheet.cell(column = j, row = i)
        cell.fill = background

save_name = input('Enter the name of the excel file without the .xlsx: ')
workbook.save(filename = f'{save_name}.xlsx')

dir_name = os.path.dirname(os.path.realpath(__file__))
print('\nYou can find the excel file in the same directory you installed the executable or ran the script in')
print(f'Location: {dir_name} \n')

delete = input('Would you like to delete the original and resized image and just keep the excel file? Y/N: ')

if delete.lower() == 'y':
    os.remove(filename)
    os.remove('input_resized.png')
    print('\nThe images have been removed. Thanks for using the script. Just run it again for the more images')
else:
    print('\nNo worries! Thanks for using the script. Just run it again for the more images')
